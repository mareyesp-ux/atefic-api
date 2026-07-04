# ============================================================
# ATEFIC API
# Análisis prospectivo de tendencias científicas con Machine Learning
# Versión con SUBCORPUS ORIENTADO POR RETO
#
# Uso local:
# 1. Crear requirements.txt
# 2. Crear entorno virtual
# 3. Instalar dependencias
# 4. Ejecutar:
#    .\.venv\Scripts\python.exe -m uvicorn app:app --reload
# 5. Abrir:
#    http://127.0.0.1:8000
# ============================================================

from pathlib import Path
import html
import re
import textwrap
import time
import unicodedata
import warnings

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from fastapi import FastAPI, Form, File, UploadFile, HTTPException, Request
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sklearn.cluster import KMeans
from sklearn.feature_extraction.text import TfidfVectorizer, ENGLISH_STOP_WORDS
from sklearn.linear_model import LinearRegression
from sklearn.metrics.pairwise import cosine_similarity

warnings.filterwarnings("ignore")


# ============================================================
# 1. RUTAS Y PARÁMETROS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"

DATA_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

ANIOS_COMPARACION = [2030, 2035, 2040]
N_CLUSTERS = 10

# Parámetros del subcorpus
SUBCORPUS_MIN_DOCS = 35
SUBCORPUS_MAX_DOCS = 160
SUBCORPUS_FRACCION = 0.28


# ============================================================
# 2. COLORES
# ============================================================

COLOR_AZUL_OSCURO = "#0B2545"
COLOR_VERDE_FIG = "#0F766E"
COLOR_NARANJA = "#F59E0B"
COLOR_GRIS_FIG = "#64748B"
COLOR_ROJO_FIG = "#DC2626"

# Colores para Excel: sin #
COLOR_HEADER = "0B2545"
COLOR_BLANCO = "FFFFFF"
COLOR_GRIS_CLARO = "F2F4F7"
COLOR_BORDE = "D9DDE3"
COLOR_TEXTO = "1F2937"
COLOR_VERDE = "D9EAD3"
COLOR_AMARILLO = "FFF2CC"
COLOR_ROJO = "F4CCCC"
COLOR_GRIS = "E7E6E6"

# Colores para Matplotlib: con #
COLOR_TEXTO_FIG = "#1F2937"

plt.rcParams["font.family"] = "DejaVu Sans"


# ============================================================
# 3. STOPWORDS
# ============================================================

STOPWORDS_ES = [
    "de", "la", "el", "los", "las", "un", "una", "unos", "unas",
    "para", "con", "por", "del", "desde", "sobre", "este", "esta",
    "estos", "estas", "en", "y", "o", "a", "que", "se", "como",
    "su", "sus", "al", "entre", "sin", "mas", "más", "menos",
    "tambien", "también", "estudio", "estudios", "analisis", "análisis",
    "investigacion", "investigación", "modelo", "modelos", "datos",
    "resultado", "resultados", "caso", "casos", "articulo", "artículo",
    "futuro", "hacia", "colombia", "colombiano", "colombiana"
]

STOPWORDS_ACADEMICAS = [
    "study", "studies", "paper", "article", "research", "analysis",
    "approach", "method", "methods", "model", "models", "result",
    "results", "evidence", "effect", "effects", "impact", "impacts",
    "role", "using", "based", "case", "cases", "new", "different",
    "data", "findings", "literature", "review", "empirical",
    "theoretical", "framework", "sample", "samples", "examines",
    "investigates", "explores", "proposes", "develops", "application"
]

STOPWORDS_EXTRA = sorted(
    set(ENGLISH_STOP_WORDS)
    .union(STOPWORDS_ES)
    .union(STOPWORDS_ACADEMICAS)
)


# ============================================================
# 4. DICCIONARIO TEMÁTICO FLEXIBLE
# ============================================================

GRUPOS_TEMATICOS = {
    "aduanero_contrabando": [
        "aduanero", "aduana", "customs", "customs control", "customs fraud",
        "contrabando", "smuggling", "contraband", "illicit trade",
        "illegal trade", "black market", "border", "border control",
        "cross border", "transfronterizo", "transfronteriza",
        "transfronterizas", "frontera", "fronterizo", "tariff",
        "arancel", "tariff classification", "hs code", "valuation",
        "customs valuation", "subfacturacion", "subfacturación",
        "misinvoicing", "trade misinvoicing", "under invoicing",
        "over invoicing", "import fraud", "export fraud", "imports",
        "exports", "redes ilicitas", "redes ilícitas", "organized crime"
    ],
    "tributario_fiscal": [
        "tributario", "tributaria", "fiscal", "tax", "taxpayer",
        "taxpayers", "tax compliance", "cumplimiento tributario",
        "evasion", "tax evasion", "evasión", "evasion fiscal",
        "evasión fiscal", "elusion", "elusión", "avoidance",
        "tax avoidance", "income tax", "corporate tax",
        "tax morale", "moral fiscal", "tax gap", "brecha tributaria",
        "informality", "informalidad", "shadow economy",
        "profit shifting", "transfer pricing", "tax haven",
        "base erosion", "fiscalizacion", "fiscalización"
    ],
    "tecnologia_digital": [
        "machine learning", "artificial intelligence", "inteligencia artificial",
        "ia", "deep learning", "neural network", "random forest",
        "algorithm", "algoritmo", "prediction", "prediccion", "predicción",
        "analytics", "analitica", "analítica", "data mining",
        "computer vision", "image detection", "x ray", "x-ray",
        "digital", "digital economy", "economia digital", "economía digital",
        "platform", "platforms", "plataformas", "digital platform",
        "e commerce", "e-commerce", "comercio electronico",
        "comercio electrónico", "online platform", "online trade",
        "blockchain", "traceability", "trazabilidad", "distributed ledger",
        "smart contract", "interoperabilidad", "data sharing",
        "information exchange", "big data"
    ],
    "riesgo_fraude": [
        "risk", "riesgo", "risk management", "gestion del riesgo",
        "gestión del riesgo", "risk assessment", "risk prediction",
        "fraud", "fraude", "fraud detection", "deteccion de fraude",
        "detección de fraude", "anomaly", "anomaly detection",
        "anomalia", "anomalía", "alert", "alerta", "profiling",
        "risk profiling", "compliance risk", "financial crime"
    ],
    "interoperabilidad_gobernanza": [
        "interoperabilidad", "interoperability", "governance",
        "gobernanza", "data governance", "gobernanza de datos",
        "data sharing", "information sharing", "information exchange",
        "institutional cooperation", "cooperacion institucional",
        "cooperación institucional", "public administration",
        "public sector", "e government", "e-government",
        "digital government", "open data", "administrative data"
    ],
    "trazabilidad_cadenas": [
        "traceability", "trazabilidad", "supply chain", "cadena de suministro",
        "cadenas de suministro", "logistics", "logistica", "logística",
        "blockchain", "distributed ledger", "smart contract",
        "product traceability", "supply chain transparency",
        "supply chain risk", "counterfeit", "counterfeiting"
    ],
    "ambiental_ilicito": [
        "wildlife", "species", "vida silvestre", "environmental crime",
        "delito ambiental", "biodiversity", "biodiversidad",
        "illegal wildlife", "wildlife trade", "illegal logging",
        "mining", "illegal mining", "natural resources"
    ],
    "productos_regulados": [
        "tobacco", "cigarette", "cigarettes", "smoking",
        "illicit tobacco", "tobacco products", "excise",
        "alcohol", "pharmaceutical", "medicine", "counterfeit medicine",
        "regulated products"
    ]
}


# ============================================================
# 5. FASTAPI
# ============================================================

app = FastAPI(
    title="ATEFIC API",
    description="Análisis prospectivo de tendencias científicas con Machine Learning.",
    version="2.0.0"
)

app.mount("/output", StaticFiles(directory=str(OUTPUT_DIR)), name="output")


# ============================================================
# 6. PÁGINA PRINCIPAL
# ============================================================

@app.get("/", response_class=HTMLResponse)
def inicio():
    return """
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <title>ATEFIC</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                background: #f3f6fb;
                margin: 0;
                padding: 0;
                color: #1f2937;
            }
            .container {
                max-width: 960px;
                margin: 40px auto;
                background: white;
                padding: 35px;
                border-radius: 18px;
                box-shadow: 0 8px 30px rgba(0,0,0,0.08);
            }
            h1 {
                color: #0B2545;
                margin-bottom: 5px;
                font-size: 38px;
            }
            .subtitulo {
                color: #64748b;
                margin-bottom: 25px;
                font-size: 18px;
            }
            label {
                font-weight: bold;
                display: block;
                margin-top: 18px;
                margin-bottom: 8px;
                font-size: 16px;
            }
            textarea {
                width: 100%;
                height: 140px;
                padding: 12px;
                border-radius: 10px;
                border: 1px solid #cbd5e1;
                font-size: 15px;
                box-sizing: border-box;
            }
            input[type="file"] {
                width: 100%;
                padding: 12px;
                border: 1px solid #cbd5e1;
                border-radius: 10px;
                background: #f8fafc;
                box-sizing: border-box;
            }
            button {
                margin-top: 25px;
                background: #0B2545;
                color: white;
                padding: 14px 28px;
                border: none;
                border-radius: 10px;
                font-size: 16px;
                font-weight: bold;
                cursor: pointer;
            }
            button:hover {
                background: #143b68;
            }
            .nota {
                margin-top: 20px;
                font-size: 13px;
                color: #64748b;
            }
            .ejemplos {
                background: #f8fafc;
                border: 1px solid #cbd5e1;
                border-radius: 12px;
                padding: 16px;
                margin-top: 20px;
                font-size: 14px;
            }
            .ejemplos strong {
                color: #0B2545;
            }
            li {
                margin-bottom: 8px;
            }
            .badge {
                display: inline-block;
                background: #D9EAD3;
                color: #274E13;
                font-weight: bold;
                padding: 6px 10px;
                border-radius: 8px;
                font-size: 13px;
                margin-bottom: 15px;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>ATEFIC</h1>
            <div class="subtitulo">
                Vigilancia científica con Machine Learning y prospectiva estratégica
            </div>

            <div class="badge">
                Versión con subcorpus orientado por reto
            </div>

            <form action="/analizar-web" method="post" enctype="multipart/form-data">
                <label>Reto prospectivo del grupo</label>
                <textarea name="reto" required>Transformación del control aduanero colombiano frente al contrabando técnico, la subfacturación comercial y las redes ilícitas transfronterizas hacia 2040</textarea>

                <label>Corpus maestro Scopus .bib</label>
                <input type="file" name="archivo_bib" accept=".bib" required>

                <button type="submit">Ejecutar análisis</button>
            </form>

            <div class="ejemplos">
                <strong>Ejemplos de retos que puedes probar:</strong>
                <ul>
                    <li>Transformación del control aduanero colombiano frente al contrabando técnico, la subfacturación comercial y las redes ilícitas transfronterizas hacia 2040.</li>
                    <li>Futuro de la fiscalización tributaria frente a la economía digital, las plataformas electrónicas y el uso de inteligencia artificial para la gestión del riesgo hacia 2040.</li>
                    <li>Futuro de la interoperabilidad institucional para fortalecer la fiscalización tributaria y aduanera mediante intercambio de información, gobernanza de datos y analítica de riesgo hacia 2035.</li>
                    <li>Futuro de la trazabilidad digital, blockchain y control de cadenas de suministro para reducir evasión, contrabando y comercio ilícito hacia 2040.</li>
                </ul>
            </div>

            <div class="nota">
                La herramienta no analiza todo el corpus de la misma forma. Primero selecciona los documentos más afines al reto del grupo y luego construye tendencias sobre ese subcorpus.
            </div>
        </div>
    </body>
    </html>
    """


# ============================================================
# 7. ENDPOINTS WEB Y API
# ============================================================

@app.post("/analizar-web", response_class=HTMLResponse)
async def analizar_web(
    request: Request,
    reto: str = Form(...),
    archivo_bib: UploadFile = File(...)
):
    try:
        reto = reto.strip()

        if len(reto) < 10:
            return pagina_error(
                "El reto es demasiado corto.",
                "Escribe un reto prospectivo más específico."
            )

        if not archivo_bib.filename.lower().endswith(".bib"):
            return pagina_error(
                "Archivo no válido.",
                "El archivo debe tener extensión .bib exportado desde Scopus."
            )

        limpiar_carpeta(DATA_DIR)
        limpiar_carpeta(OUTPUT_DIR)

        destino = DATA_DIR / "archivo_usuario.bib"
        contenido = await archivo_bib.read()

        with open(destino, "wb") as f:
            f.write(contenido)

        inicio_tiempo = time.time()
        resultado = ejecutar_analisis(reto)
        duracion = round(time.time() - inicio_tiempo, 2)

        base_url = str(request.base_url).rstrip("/")
        archivos = listar_archivos(base_url)

        enlaces_html = ""
        imagenes_html = ""

        for archivo in archivos:
            enlaces_html += f"""
            <li>
                <a href="{archivo['url']}" target="_blank">{archivo['nombre']}</a>
            </li>
            """

            if archivo["nombre"].lower().endswith(".png"):
                imagenes_html += f"""
                <h4>{archivo['nombre']}</h4>
                <img src="{archivo['url']}" style="width:100%; border:1px solid #cbd5e1; border-radius:12px; margin-bottom:25px;">
                """

        return f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <title>ATEFIC · Resultados</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    background: #f3f6fb;
                    color: #1f2937;
                }}
                .container {{
                    max-width: 1080px;
                    margin: 40px auto;
                    background: white;
                    padding: 35px;
                    border-radius: 18px;
                    box-shadow: 0 8px 30px rgba(0,0,0,0.08);
                }}
                h1 {{
                    color: #0B2545;
                }}
                .ok {{
                    background: #D9EAD3;
                    padding: 15px;
                    border-radius: 10px;
                    font-weight: bold;
                    color: #274E13;
                }}
                a {{
                    color: #0B2545;
                    font-weight: bold;
                }}
                li {{
                    margin-bottom: 10px;
                }}
                .reto {{
                    background: #f8fafc;
                    padding: 14px;
                    border-radius: 10px;
                    border: 1px solid #cbd5e1;
                    margin-top: 12px;
                }}
                pre {{
                    white-space: pre-wrap;
                    background: #f8fafc;
                    padding: 15px;
                    border-radius: 10px;
                    border: 1px solid #cbd5e1;
                    font-size: 13px;
                }}
                .volver {{
                    display: inline-block;
                    margin-top: 25px;
                    background: #0B2545;
                    color: white;
                    padding: 12px 22px;
                    border-radius: 10px;
                    text-decoration: none;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>ATEFIC · Resultados generados</h1>

                <div class="ok">
                    Análisis ejecutado correctamente en {duracion} segundos.
                </div>

                <h3>Reto analizado</h3>
                <div class="reto">{html.escape(reto)}</div>

                <h3>Archivos descargables</h3>
                <ul>
                    {enlaces_html}
                </ul>

                <h3>Vista previa del análisis</h3>
                <pre>{html.escape(resultado)}</pre>

                <h3>Figuras generadas</h3>
                {imagenes_html}

                <a class="volver" href="/">Ejecutar otro análisis</a>
            </div>
        </body>
        </html>
        """

    except Exception as e:
        return pagina_error("Ocurrió un error durante el análisis.", str(e))


@app.post("/analizar")
async def analizar_api(
    reto: str = Form(...),
    archivo_bib: UploadFile = File(...)
):
    try:
        reto = reto.strip()

        limpiar_carpeta(DATA_DIR)
        limpiar_carpeta(OUTPUT_DIR)

        destino = DATA_DIR / "archivo_usuario.bib"
        contenido = await archivo_bib.read()

        with open(destino, "wb") as f:
            f.write(contenido)

        resumen = ejecutar_analisis(reto)

        return {
            "status": "ok",
            "reto": reto,
            "archivos": [archivo.name for archivo in OUTPUT_DIR.glob("*") if archivo.is_file()],
            "resumen": resumen
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/health")
def health():
    return {
        "status": "ok",
        "version": "2.0.0-subcorpus-orientado-por-reto",
        "data_dir": str(DATA_DIR),
        "output_dir": str(OUTPUT_DIR)
    }


def pagina_error(titulo, detalle):
    return f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <title>ATEFIC · Error</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                background: #f3f6fb;
                color: #1f2937;
            }}
            .container {{
                max-width: 900px;
                margin: 40px auto;
                background: white;
                padding: 35px;
                border-radius: 18px;
                box-shadow: 0 8px 30px rgba(0,0,0,0.08);
            }}
            h1 {{
                color: #7F0000;
            }}
            .error {{
                background: #F4CCCC;
                padding: 15px;
                border-radius: 10px;
                color: #7F0000;
                font-weight: bold;
            }}
            pre {{
                white-space: pre-wrap;
                background: #f8fafc;
                padding: 15px;
                border-radius: 10px;
                border: 1px solid #cbd5e1;
                font-size: 13px;
            }}
            a {{
                display: inline-block;
                margin-top: 25px;
                background: #0B2545;
                color: white;
                padding: 12px 22px;
                border-radius: 10px;
                text-decoration: none;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>ATEFIC · Error</h1>
            <div class="error">{html.escape(titulo)}</div>
            <h3>Detalle técnico</h3>
            <pre>{html.escape(detalle)}</pre>
            <a href="/">Volver al formulario</a>
        </div>
    </body>
    </html>
    """


# ============================================================
# 8. UTILIDADES
# ============================================================

def limpiar_carpeta(carpeta):
    carpeta.mkdir(exist_ok=True)

    for archivo in carpeta.glob("*"):
        if archivo.is_file():
            try:
                archivo.unlink()
            except Exception:
                pass


def listar_archivos(base_url):
    archivos = []

    for archivo in sorted(OUTPUT_DIR.glob("*")):
        if archivo.is_file():
            archivos.append({
                "nombre": archivo.name,
                "url": f"{base_url}/output/{archivo.name}"
            })

    return archivos


def quitar_tildes(texto):
    texto = str(texto)
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return texto


def normalizar(texto):
    texto = quitar_tildes(str(texto).lower())
    texto = re.sub(r"[^a-z0-9\s\-]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def envolver_texto(texto, ancho=38):
    return "\n".join(textwrap.wrap(str(texto), width=ancho))


def recortar_texto(texto, max_caracteres=95):
    texto = str(texto)

    if len(texto) <= max_caracteres:
        return texto

    return texto[:max_caracteres - 3].rstrip() + "..."


# ============================================================
# 9. LECTURA BIBTEX
# ============================================================

def limpiar_campo(valor):
    if valor is None:
        return ""

    valor = str(valor)
    valor = valor.replace("{", " ").replace("}", " ")
    valor = valor.replace("\n", " ").replace("\r", " ")
    valor = re.sub(r"\s+", " ", valor).strip()

    return valor


def extraer_campo_bibtex(registro, campo):
    patron = (
        rf"\b{campo}\s*=\s*"
        rf"[\{{\"]"
        rf"(.*?)"
        rf"[\}}\"]\s*,?\s*(?=\n\s*[A-Za-z_]+\s*=|\n\s*\}})"
    )

    match = re.search(patron, registro, flags=re.IGNORECASE | re.DOTALL)

    if match:
        return limpiar_campo(match.group(1))

    return ""


def leer_bibtex():
    archivos = sorted(DATA_DIR.glob("*.bib"))

    if not archivos:
        raise ValueError("No hay archivo .bib cargado.")

    ruta_bib = archivos[0]

    with open(ruta_bib, "r", encoding="utf-8", errors="ignore") as f:
        contenido = f.read()

    primer_registro = contenido.find("@")

    if primer_registro == -1:
        raise ValueError("El archivo .bib no contiene registros BibTeX válidos.")

    contenido = contenido[primer_registro:]
    registros_crudos = re.split(r"\n(?=@)", contenido)

    campos = [
        "title",
        "abstract",
        "keywords",
        "author_keywords",
        "index_keywords",
        "year",
        "author",
        "journal",
        "booktitle",
        "source",
        "doi"
    ]

    filas = []

    for i, registro in enumerate(registros_crudos, start=1):
        if not registro.strip().startswith("@"):
            continue

        entry = {}

        for campo in campos:
            entry[campo] = extraer_campo_bibtex(registro, campo)

        titulo = entry.get("title", "")
        abstract = entry.get("abstract", "")
        keywords = (
            entry.get("keywords", "")
            or entry.get("author_keywords", "")
            or entry.get("index_keywords", "")
        )

        texto = f"{titulo}. {abstract}. {keywords}"

        filas.append({
            "id": i,
            "titulo": titulo,
            "abstract": abstract,
            "keywords": keywords,
            "year": entry.get("year", ""),
            "autores": entry.get("author", ""),
            "revista": entry.get("journal", "") or entry.get("booktitle", "") or entry.get("source", ""),
            "doi": entry.get("doi", ""),
            "texto": texto
        })

    df = pd.DataFrame(filas)

    if df.empty:
        raise ValueError("No se pudieron leer registros útiles del archivo .bib.")

    df["year_num"] = pd.to_numeric(df["year"], errors="coerce")

    df["texto_limpio"] = df["texto"].fillna("").astype(str).str.lower()
    df["texto_limpio"] = df["texto_limpio"].apply(
        lambda x: re.sub(r"[^a-záéíóúñü0-9\s\-]", " ", x)
    )
    df["texto_limpio"] = df["texto_limpio"].apply(
        lambda x: re.sub(r"\s+", " ", x).strip()
    )

    df = df[df["texto_limpio"].str.len() > 10].copy()

    if df.empty:
        raise ValueError("El archivo no tiene texto suficiente para analizar.")

    return df, ruta_bib


# ============================================================
# 10. SUBCORPUS ORIENTADO POR RETO
# ============================================================

def presencia_terminos(texto, terminos):
    texto = normalizar(texto)

    if not texto:
        return 0.0

    conteo = 0

    for termino in terminos:
        termino_norm = normalizar(termino)

        if not termino_norm:
            continue

        if termino_norm in texto:
            conteo += texto.count(termino_norm)

    return min(1.0, conteo / 4)


def tokens_relevantes(texto):
    texto = normalizar(texto)

    stopwords = set(STOPWORDS_ES).union(STOPWORDS_ACADEMICAS).union(ENGLISH_STOP_WORDS)

    return set(
        t for t in texto.split()
        if len(t) >= 4 and t not in stopwords
    )


def perfil_reto(reto):
    r = normalizar(reto)

    activados = {}

    for nombre, terminos in GRUPOS_TEMATICOS.items():
        activados[nombre] = presencia_terminos(r, terminos)

    total = sum(activados.values())

    if total == 0:
        return {k: 1 / len(GRUPOS_TEMATICOS) for k in GRUPOS_TEMATICOS.keys()}

    return {k: v / total for k, v in activados.items()}


def construir_consulta_expandida(reto):
    """
    Convierte el reto del grupo en una consulta semántica ampliada.
    No necesita conocer previamente el reto.
    Usa el texto del reto y activa vocabulario relacionado.
    """

    reto_norm = normalizar(reto)
    pesos = perfil_reto(reto_norm)

    terminos_extra = []

    for grupo, peso in pesos.items():
        if peso >= 0.18:
            terminos_extra.extend(GRUPOS_TEMATICOS.get(grupo, [])[:22])

    consulta = reto_norm + " " + " ".join(terminos_extra)

    return consulta.strip()


def calcular_score_documento_reto(df, reto):
    """
    Calcula la afinidad de cada documento con el reto.
    Combina:
    1. Similitud TF-IDF entre reto expandido y documento.
    2. Cobertura de términos relevantes.
    3. Coincidencia del perfil temático del reto.
    """

    consulta = construir_consulta_expandida(reto)
    textos = df["texto_limpio"].fillna("").astype(str).tolist()

    try:
        vectorizer = TfidfVectorizer(
            stop_words=STOPWORDS_EXTRA,
            ngram_range=(1, 3),
            min_df=1,
            max_features=7000
        )

        matriz = vectorizer.fit_transform([consulta] + textos)
        similitudes = cosine_similarity(matriz[0], matriz[1:]).flatten()

    except Exception:
        similitudes = np.zeros(len(df))

    tokens_consulta = tokens_relevantes(consulta)
    pesos_reto = perfil_reto(reto)

    coberturas = []
    perfiles = []

    for texto in textos:
        tokens_doc = tokens_relevantes(texto)

        if tokens_consulta and tokens_doc:
            cobertura = len(tokens_consulta.intersection(tokens_doc)) / max(len(tokens_consulta), 1)
        else:
            cobertura = 0.0

        score_perfil = 0.0

        for grupo, peso in pesos_reto.items():
            score_perfil += peso * presencia_terminos(texto, GRUPOS_TEMATICOS.get(grupo, []))

        coberturas.append(cobertura)
        perfiles.append(score_perfil)

    similitudes = np.array(similitudes)
    coberturas = np.array(coberturas)
    perfiles = np.array(perfiles)

    score = (
        0.52 * similitudes
        + 0.20 * coberturas
        + 0.28 * perfiles
    )

    return score, similitudes, coberturas, perfiles, consulta


def filtrar_documentos_por_reto(df, reto):
    """
    Construye el subcorpus de análisis para cada grupo.
    Este es el cambio clave: el clustering ya no se hace sobre todo el .bib,
    sino sobre los documentos más relacionados con el reto escrito.
    """

    df = df.copy()
    n_total = len(df)

    score, similitud, cobertura, perfil, consulta = calcular_score_documento_reto(df, reto)

    df["score_reto"] = score
    df["similitud_reto"] = similitud
    df["cobertura_reto"] = cobertura
    df["perfil_reto_score"] = perfil

    df = df.sort_values("score_reto", ascending=False).reset_index(drop=True)

    if n_total <= SUBCORPUS_MIN_DOCS:
        top_k = n_total
    else:
        top_k = int(n_total * SUBCORPUS_FRACCION)
        top_k = max(SUBCORPUS_MIN_DOCS, top_k)
        top_k = min(SUBCORPUS_MAX_DOCS, top_k)
        top_k = min(n_total, top_k)

    subcorpus = df.head(top_k).copy()

    # Si el corpus es grande, elimina documentos con afinidad casi nula,
    # siempre que no destruya el mínimo necesario.
    if n_total > SUBCORPUS_MIN_DOCS:
        filtrado_positivo = subcorpus[subcorpus["score_reto"] > 0.01].copy()

        if len(filtrado_positivo) >= max(15, min(SUBCORPUS_MIN_DOCS, top_k)):
            subcorpus = filtrado_positivo

    info = {
        "documentos_corpus_maestro": int(n_total),
        "documentos_subcorpus": int(len(subcorpus)),
        "porcentaje_subcorpus": round((len(subcorpus) / n_total) * 100, 2),
        "score_max": round(float(subcorpus["score_reto"].max()), 4) if len(subcorpus) > 0 else 0,
        "score_min": round(float(subcorpus["score_reto"].min()), 4) if len(subcorpus) > 0 else 0,
        "score_promedio": round(float(subcorpus["score_reto"].mean()), 4) if len(subcorpus) > 0 else 0,
        "consulta_expandida": consulta
    }

    if len(subcorpus) < 2:
        raise ValueError(
            "El reto produjo un subcorpus demasiado pequeño. "
            "Prueba con un reto más amplio o carga un corpus maestro más grande."
        )

    return subcorpus, info


# ============================================================
# 11. AFINIDAD DE CLÚSTER CON RETO
# ============================================================

def afinidad_con_reto(reto, texto_tendencia, texto_cluster=""):
    """
    Afinidad del clúster con el reto.
    Aquí ya trabajamos sobre un subcorpus filtrado,
    pero esta afinidad ayuda a ordenar y calcular IRP.
    """

    reto_norm = normalizar(reto)
    tendencia_norm = normalizar(texto_tendencia)
    cluster_norm = normalizar(texto_cluster[:12000])

    texto_base = f"{tendencia_norm} {tendencia_norm} {tendencia_norm} {cluster_norm}"

    pesos_reto = perfil_reto(reto_norm)

    puntaje_perfil = 0.0

    for grupo, peso in pesos_reto.items():
        puntaje_perfil += peso * presencia_terminos(texto_base, GRUPOS_TEMATICOS.get(grupo, []))

    try:
        consulta = construir_consulta_expandida(reto_norm)

        vectorizer = TfidfVectorizer(
            stop_words=STOPWORDS_EXTRA,
            ngram_range=(1, 3),
            min_df=1,
            max_features=3000
        )

        matriz = vectorizer.fit_transform([consulta, texto_base])
        sim_tfidf = float(cosine_similarity(matriz[0], matriz[1])[0][0])

    except Exception:
        sim_tfidf = 0.0

    tokens_reto = tokens_relevantes(construir_consulta_expandida(reto_norm))
    tokens_texto = tokens_relevantes(texto_base)

    if tokens_reto and tokens_texto:
        cobertura_reto = len(tokens_reto.intersection(tokens_texto)) / max(len(tokens_reto), 1)
    else:
        cobertura_reto = 0.0

    afinidad = (
        0.10
        + 0.60 * puntaje_perfil
        + 0.25 * sim_tfidf
        + 0.05 * cobertura_reto
    )

    return round(max(0.05, min(0.98, afinidad)), 3)


# ============================================================
# 12. NOMBRES DE TENDENCIAS
# ============================================================

REGLAS_TEMAS = [
    {
        "nombre": "Evasión fiscal, cumplimiento tributario y moral fiscal",
        "palabras": [
            "tax morale", "taxpayer", "taxpayers", "income tax",
            "tax compliance", "tax evasion", "tax gap", "informality",
            "shadow economy", "compliance", "evasion", "evasión",
            "cumplimiento tributario", "moral fiscal"
        ]
    },
    {
        "nombre": "Elusión fiscal corporativa y planeación tributaria agresiva",
        "palabras": [
            "tax avoidance", "corporate avoidance", "corporate tax",
            "profit shifting", "transfer pricing", "tax haven",
            "base erosion", "multinational", "corporate", "avoidance",
            "income shifting", "elusión", "planeación tributaria"
        ]
    },
    {
        "nombre": "IA, machine learning y detección automatizada de fraude",
        "palabras": [
            "machine learning", "artificial intelligence", "deep learning",
            "random forest", "neural network", "fraud detection",
            "anomaly detection", "risk prediction", "x ray", "x-ray",
            "image detection", "algorithm", "classification", "prediction",
            "learning", "computer vision", "inteligencia artificial",
            "analítica", "analitica"
        ]
    },
    {
        "nombre": "Fraude aduanero, clasificación arancelaria y subfacturación",
        "palabras": [
            "customs fraud", "customs valuation", "misinvoicing",
            "trade misinvoicing", "under invoicing", "over invoicing",
            "tariff classification", "hs code", "misclassified", "imports",
            "import fraud", "export fraud", "tariff", "customs", "valuation",
            "subfacturacion", "subfacturación", "arancel"
        ]
    },
    {
        "nombre": "Contrabando, comercio ilícito y mercados ilegales",
        "palabras": [
            "smuggling", "contraband", "illicit trade", "illegal trade",
            "black market", "illicit market", "organized crime",
            "cross border", "border control", "illegal", "illicit", "crime",
            "contrabando", "frontera", "transfronterizo"
        ]
    },
    {
        "nombre": "Contrabando de tabaco, cigarrillos y productos regulados",
        "palabras": [
            "tobacco", "cigarette", "cigarettes", "smoking",
            "illicit tobacco", "tobacco products", "excise"
        ]
    },
    {
        "nombre": "Tráfico de especies, vida silvestre y comercio ilegal ambiental",
        "palabras": [
            "wildlife", "species", "conservation", "illegal wildlife",
            "wildlife trade", "environmental crime", "biodiversity",
            "vida silvestre", "delito ambiental"
        ]
    },
    {
        "nombre": "Trazabilidad digital, blockchain y cadenas de suministro",
        "palabras": [
            "blockchain", "traceability", "supply chain",
            "distributed ledger", "smart contract", "digital traceability",
            "trazabilidad", "cadenas de suministro"
        ]
    },
    {
        "nombre": "Economía digital, plataformas y nuevos retos de fiscalización",
        "palabras": [
            "digital economy", "e commerce", "e-commerce", "platform economy",
            "digital platform", "online platform", "cross border e commerce",
            "economía digital", "economia digital", "plataformas"
        ]
    },
    {
        "nombre": "Gobernanza de datos, interoperabilidad y cooperación institucional",
        "palabras": [
            "data sharing", "interoperability", "information exchange",
            "governance", "institutional cooperation", "risk management",
            "interoperabilidad", "gobernanza", "cooperación institucional"
        ]
    }
]


def nombre_generico_cluster(terminos_top):
    limpios = []

    for termino in terminos_top:
        termino = str(termino).strip()

        if len(termino) < 3:
            continue

        if termino.lower() in STOPWORDS_EXTRA:
            continue

        limpios.append(termino)

    if not limpios:
        return "Tendencia científica emergente identificada en el subcorpus"

    return "Tendencia científica asociada a: " + ", ".join(limpios[:5])


def nombrar_cluster(texto_cluster, terminos_top):
    top = " ".join(terminos_top).lower()
    muestra = texto_cluster.lower()[:70000]
    combinado = top + " " + muestra

    mejor_nombre = None
    mejor_score = 0

    for regla in REGLAS_TEMAS:
        score = 0

        for palabra in regla["palabras"]:
            palabra = palabra.lower()

            if palabra in top:
                score += 10

            apariciones = combinado.count(palabra)

            if apariciones > 0:
                score += min(apariciones, 15)

        if score > mejor_score:
            mejor_score = score
            mejor_nombre = regla["nombre"]

    if mejor_nombre and mejor_score >= 3:
        return mejor_nombre

    return nombre_generico_cluster(terminos_top)


def hacer_nombre_unico(base, terminos_top, usados):
    if base not in usados:
        usados[base] = 1
        return base

    usados[base] += 1

    subtema = ", ".join([t for t in terminos_top[:4] if len(t) > 2])

    return f"{base} · Subtema: {subtema}"


# ============================================================
# 13. CLASIFICACIONES
# ============================================================

def clasificar_madurez(n_docs, recent_share, total_docs):
    if n_docs >= max(40, total_docs * 0.18):
        return "Tendencia consolidada en el subcorpus"

    if n_docs >= 15 and recent_share >= 0.42:
        return "Tendencia emergente"

    if n_docs < 15 and recent_share >= 0.42:
        return "Señal débil"

    if n_docs >= 15:
        return "Tema de monitoreo"

    return "Hipótesis a validar"


def clasificar_relevancia_prospectiva(irp):
    if irp >= 0.78:
        return "Prioridad prospectiva crítica"

    if irp >= 0.62:
        return "Prioridad prospectiva alta"

    if irp >= 0.48:
        return "Señal prospectiva consolidada"

    return "Señal débil en vigilancia"


def etiqueta_corta(irp):
    if irp >= 0.78:
        return "Crítica"

    if irp >= 0.62:
        return "Alta"

    if irp >= 0.48:
        return "Consolidada"

    return "Vigilancia"


def impacto_desde_irp(irp):
    if irp >= 0.78:
        return "Muy alto"

    if irp >= 0.62:
        return "Alto"

    if irp >= 0.48:
        return "Medio-alto"

    return "Medio"


def velocidad_desde_recencia(proporcion_reciente):
    if proporcion_reciente >= 0.60:
        return "Rápida"

    if proporcion_reciente >= 0.35:
        return "Moderada"

    return "Lenta o no estimable"


# ============================================================
# 14. MACHINE LEARNING
# ============================================================

def construir_vectorizador(textos):
    intentos = [
        {"min_df": 3, "max_df": 0.88},
        {"min_df": 2, "max_df": 0.95},
        {"min_df": 1, "max_df": 1.0},
    ]

    ultimo_error = None

    for params in intentos:
        try:
            vectorizer = TfidfVectorizer(
                max_features=4500,
                min_df=params["min_df"],
                max_df=params["max_df"],
                ngram_range=(1, 3),
                stop_words=STOPWORDS_EXTRA
            )

            matriz = vectorizer.fit_transform(textos)

            if matriz.shape[1] > 0:
                return vectorizer, matriz, params

        except Exception as e:
            ultimo_error = e

    raise ValueError(f"No fue posible construir la matriz TF-IDF. Error: {ultimo_error}")


def construir_clusters(df, reto):
    vectorizer, matriz_tfidf, params = construir_vectorizador(df["texto_limpio"])

    n_docs = len(df)

    if n_docs == 1:
        df["cluster"] = 0
        terminos = np.array(vectorizer.get_feature_names_out()).tolist()
        texto_cluster = df["texto_limpio"].iloc[0]
        nombre = nombre_generico_cluster(terminos[:10])

        clusters = [{
            "cluster": 0,
            "tendencia": nombre,
            "documentos": 1,
            "porcentaje_corpus": 100.0,
            "terminos_dominantes": terminos[:10],
            "documentos_recientes": 1,
            "proporcion_reciente": 1.0,
            "madurez": "Hipótesis a validar",
            "texto_cluster": texto_cluster,
            "afinidad_reto": afinidad_con_reto(reto, nombre, texto_cluster),
            "score_documental_promedio": float(df["score_reto"].mean()) if "score_reto" in df.columns else 0
        }]

        parametros = {
            "total_documentos": 1,
            "numero_clusters": 1,
            "min_df": params["min_df"],
            "max_df": params["max_df"]
        }

        return df, clusters, parametros

    n_clusters = max(2, min(N_CLUSTERS, n_docs))

    modelo = KMeans(
        n_clusters=n_clusters,
        random_state=42,
        n_init=20
    )

    df["cluster"] = modelo.fit_predict(matriz_tfidf)

    terminos = np.array(vectorizer.get_feature_names_out())

    max_year = df["year_num"].dropna().max()
    recent_from = None if pd.isna(max_year) else int(max_year) - 3

    total_docs = len(df)
    clusters = []
    usados = {}

    for cluster_id in sorted(df["cluster"].unique()):
        sub = df[df["cluster"] == cluster_id].copy()

        centro = modelo.cluster_centers_[cluster_id]
        top_idx = centro.argsort()[::-1][:18]
        terminos_top = terminos[top_idx].tolist()

        texto_cluster = " ".join(sub["texto_limpio"].tolist())

        nombre_base = nombrar_cluster(texto_cluster, terminos_top)
        nombre_unico = hacer_nombre_unico(nombre_base, terminos_top, usados)

        n_docs_cluster = len(sub)

        if recent_from is not None:
            con_anio = sub.dropna(subset=["year_num"])
            if len(con_anio) > 0:
                recent_count = int((con_anio["year_num"] >= recent_from).sum())
                recent_share = recent_count / len(con_anio)
            else:
                recent_count = 0
                recent_share = 0
        else:
            recent_count = 0
            recent_share = 0

        if "score_reto" in sub.columns:
            score_documental_promedio = round(float(sub["score_reto"].mean()), 4)
        else:
            score_documental_promedio = 0.0

        clusters.append({
            "cluster": int(cluster_id),
            "tendencia": nombre_unico,
            "documentos": int(n_docs_cluster),
            "porcentaje_corpus": round((n_docs_cluster / total_docs) * 100, 2),
            "terminos_dominantes": terminos_top[:10],
            "documentos_recientes": int(recent_count),
            "proporcion_reciente": round(float(recent_share), 3),
            "madurez": clasificar_madurez(n_docs_cluster, recent_share, total_docs),
            "texto_cluster": texto_cluster,
            "afinidad_reto": afinidad_con_reto(reto, nombre_unico, texto_cluster),
            "score_documental_promedio": score_documental_promedio
        })

    parametros = {
        "total_documentos": total_docs,
        "numero_clusters": n_clusters,
        "min_df": params["min_df"],
        "max_df": params["max_df"]
    }

    return df, clusters, parametros


def construir_proyeccion(df, clusters):
    anio_min = df["year_num"].dropna().min()
    anio_max = df["year_num"].dropna().max()

    if pd.isna(anio_min) or pd.isna(anio_max):
        anio_min = 2020
        anio_max = 2026
    else:
        anio_min = int(anio_min)
        anio_max = int(anio_max)

    if anio_min == anio_max:
        anio_min = anio_min - 1

    filas = []

    for c in clusters:
        cluster_id = c["cluster"]

        sub = df[df["cluster"] == cluster_id].dropna(subset=["year_num"]).copy()

        conteo = sub.groupby("year_num").size().reset_index(name="documentos")

        if conteo.empty:
            conteo = pd.DataFrame({
                "year_num": [anio_min, anio_max],
                "documentos": [0, c["documentos"]]
            })

        conteo["year_num"] = conteo["year_num"].astype(int)

        base_anios = pd.DataFrame({
            "year_num": list(range(anio_min, anio_max + 1))
        })

        serie = base_anios.merge(conteo, on="year_num", how="left")
        serie["documentos"] = serie["documentos"].fillna(0)

        X = serie[["year_num"]].values
        y = np.log1p(serie["documentos"].values)

        if len(serie) >= 2 and serie["documentos"].sum() > 0:
            reg = LinearRegression().fit(X, y)
            pendiente = float(reg.coef_[0])
            r2 = float(reg.score(X, y))
        else:
            reg = None
            pendiente = 0.0
            r2 = 0.0

        for anio in ANIOS_COMPARACION:
            if reg is not None:
                pred_log = reg.predict(np.array([[anio]]))[0]
                docs_proyectados = max(0.0, float(np.expm1(pred_log)))
            else:
                docs_proyectados = 0.0

            filas.append({
                "anio": anio,
                "cluster": cluster_id,
                "tendencia": c["tendencia"],
                "documentos_proyectados": docs_proyectados,
                "pendiente_crecimiento": pendiente,
                "r2_tendencia": r2,
                "proporcion_reciente": c["proporcion_reciente"],
                "afinidad_reto": c["afinidad_reto"],
                "score_documental_promedio": c.get("score_documental_promedio", 0.0),
            })

    df_pred = pd.DataFrame(filas)
    resultados = []

    for anio in ANIOS_COMPARACION:
        sub = df_pred[df_pred["anio"] == anio].copy()

        min_docs = sub["documentos_proyectados"].min()
        max_docs = sub["documentos_proyectados"].max()

        min_pend = sub["pendiente_crecimiento"].min()
        max_pend = sub["pendiente_crecimiento"].max()

        min_doc_score = sub["score_documental_promedio"].min()
        max_doc_score = sub["score_documental_promedio"].max()

        def normalizar_valor(valor, minimo, maximo):
            if maximo == minimo:
                return 0.5
            return (valor - minimo) / (maximo - minimo)

        for _, row in sub.iterrows():
            docs_norm = normalizar_valor(row["documentos_proyectados"], min_docs, max_docs)
            pend_norm = normalizar_valor(row["pendiente_crecimiento"], min_pend, max_pend)
            score_doc_norm = normalizar_valor(row["score_documental_promedio"], min_doc_score, max_doc_score)

            if anio == 2030:
                irp = (
                    0.14 * docs_norm
                    + 0.10 * pend_norm
                    + 0.06 * row["proporcion_reciente"]
                    + 0.55 * row["afinidad_reto"]
                    + 0.15 * score_doc_norm
                )
            elif anio == 2035:
                irp = (
                    0.10 * docs_norm
                    + 0.08 * pend_norm
                    + 0.05 * row["proporcion_reciente"]
                    + 0.62 * row["afinidad_reto"]
                    + 0.15 * score_doc_norm
                )
            else:
                irp = (
                    0.08 * docs_norm
                    + 0.06 * pend_norm
                    + 0.04 * row["proporcion_reciente"]
                    + 0.67 * row["afinidad_reto"]
                    + 0.15 * score_doc_norm
                )

            resultados.append({
                "anio": anio,
                "cluster": int(row["cluster"]),
                "tendencia": row["tendencia"],
                "irp": round(float(irp), 3),
                "clasificacion": clasificar_relevancia_prospectiva(irp),
                "etiqueta": etiqueta_corta(irp),
                "documentos_proyectados": round(float(row["documentos_proyectados"]), 3),
                "pendiente_crecimiento": round(float(row["pendiente_crecimiento"]), 5),
                "r2_tendencia": round(float(row["r2_tendencia"]), 3),
                "proporcion_reciente": round(float(row["proporcion_reciente"]), 3),
                "afinidad_reto": round(float(row["afinidad_reto"]), 3),
                "score_documental_promedio": round(float(row["score_documental_promedio"]), 4),
            })

    return pd.DataFrame(resultados)


# ============================================================
# 15. TABLAS
# ============================================================

def construir_tabla_tendencias(clusters, df_scores):
    s2030 = df_scores[df_scores["anio"] == 2030].set_index("cluster").to_dict(orient="index")
    s2035 = df_scores[df_scores["anio"] == 2035].set_index("cluster").to_dict(orient="index")
    s2040 = df_scores[df_scores["anio"] == 2040].set_index("cluster").to_dict(orient="index")

    filas = []

    for c in clusters:
        cluster_id = c["cluster"]

        a = s2030.get(cluster_id, {})
        b = s2035.get(cluster_id, {})
        d = s2040.get(cluster_id, {})

        irp_2030 = float(a.get("irp", 0))
        irp_2035 = float(b.get("irp", 0))
        irp_2040 = float(d.get("irp", 0))

        evidencia = (
            f"{c['documentos']} documentos asociados en el subcorpus "
            f"({c['porcentaje_corpus']}% del subcorpus). "
            f"Términos dominantes: {', '.join(c['terminos_dominantes'][:6])}."
        )

        orden = (
            0.50 * c["afinidad_reto"]
            + 0.20 * c.get("score_documental_promedio", 0)
            + 0.10 * irp_2030
            + 0.20 * irp_2040
        )

        filas.append({
            "Tendencia identificada": c["tendencia"],
            "Evidencia o señal que la sustenta": evidencia,
            "Nivel de impacto": impacto_desde_irp(irp_2030),
            "Velocidad de evolución": velocidad_desde_recencia(c["proporcion_reciente"]),
            "Madurez de la señal": c["madurez"],
            "Afinidad con el reto": c["afinidad_reto"],
            "Score documental del subcorpus": c.get("score_documental_promedio", 0),
            "2030": f"{a.get('clasificacion', 'ND')} (IRP={irp_2030:.2f})",
            "2035": f"{b.get('clasificacion', 'ND')} (IRP={irp_2035:.2f})",
            "2040": f"{d.get('clasificacion', 'ND')} (IRP={irp_2040:.2f})",
            "_orden": orden
        })

    tabla = pd.DataFrame(filas)
    tabla = tabla.sort_values("_orden", ascending=False).drop(columns=["_orden"])

    return tabla


def construir_tabla_subcorpus(df_sub):
    columnas = [
        "titulo",
        "year",
        "revista",
        "autores",
        "doi",
        "score_reto",
        "similitud_reto",
        "cobertura_reto",
        "perfil_reto_score"
    ]

    existentes = [c for c in columnas if c in df_sub.columns]

    tabla = df_sub[existentes].copy()

    renombrar = {
        "titulo": "Título",
        "year": "Año",
        "revista": "Fuente",
        "autores": "Autores",
        "doi": "DOI",
        "score_reto": "Score de afinidad con el reto",
        "similitud_reto": "Similitud TF-IDF",
        "cobertura_reto": "Cobertura de términos del reto",
        "perfil_reto_score": "Coincidencia de perfil temático"
    }

    tabla = tabla.rename(columns=renombrar)

    for col in [
        "Score de afinidad con el reto",
        "Similitud TF-IDF",
        "Cobertura de términos del reto",
        "Coincidencia de perfil temático"
    ]:
        if col in tabla.columns:
            tabla[col] = tabla[col].round(4)

    return tabla


# ============================================================
# 16. EXCEL
# ============================================================

def exportar_excel(tabla_tendencias, tabla_subcorpus, reto, ruta_bib, parametros, info_subcorpus):
    ruta_excel = OUTPUT_DIR / "ATEFIC_Tabla_Tendencias_ML.xlsx"

    parametros_df = pd.DataFrame([
        ["Reto analizado", reto],
        ["Archivo fuente", str(ruta_bib)],
        ["Documentos del corpus maestro", info_subcorpus["documentos_corpus_maestro"]],
        ["Documentos seleccionados en el subcorpus", info_subcorpus["documentos_subcorpus"]],
        ["Porcentaje del corpus usado", f"{info_subcorpus['porcentaje_subcorpus']}%"],
        ["Score máximo del subcorpus", info_subcorpus["score_max"]],
        ["Score mínimo del subcorpus", info_subcorpus["score_min"]],
        ["Score promedio del subcorpus", info_subcorpus["score_promedio"]],
        ["Número de clústeres", parametros["numero_clusters"]],
        ["Años de previsión", "2030, 2035, 2040"],
        ["Modelo textual", "TF-IDF"],
        ["Modelo de agrupamiento", "K-Means"],
        ["Indicador", "IRP - Índice de Relevancia Prospectiva"],
        ["Método clave", "Subcorpus orientado por reto"],
        ["Criterio de ajuste al reto", "Selección automática de documentos afines al reto del grupo antes del clustering"],
        ["Consulta semántica expandida", info_subcorpus["consulta_expandida"]],
        ["Nota", "El resultado se basa en el corpus científico cargado. No constituye predicción determinista."]
    ], columns=["Parámetro", "Valor"])

    with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
        tabla_tendencias.to_excel(writer, sheet_name="Tendencias", index=False, startrow=3)
        tabla_subcorpus.to_excel(writer, sheet_name="Subcorpus", index=False, startrow=3)
        parametros_df.to_excel(writer, sheet_name="Parametros", index=False, startrow=3)

        for nombre_hoja in ["Tendencias", "Subcorpus", "Parametros"]:
            ws = writer.book[nombre_hoja]
            estilizar_hoja(ws, nombre_hoja)

    return ruta_excel


def estilizar_hoja(ws, nombre_hoja):
    ws.sheet_view.showGridLines = False

    max_col = ws.max_column
    final_col = get_column_letter(max_col)

    ws.merge_cells(f"A1:{final_col}1")
    ws["A1"] = f"ATEFIC · {nombre_hoja}"
    ws["A1"].font = Font(name="Cambria", size=16, bold=True, color=COLOR_AZUL_OSCURO.replace("#", ""))
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{final_col}2")
    ws["A2"] = "Vigilancia científica con Machine Learning y subcorpus orientado por reto"
    ws["A2"].font = Font(name="Cambria", size=11, italic=True, color="5B6472")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30

    thin = Side(style="thin", color=COLOR_BORDE)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_row = 4

    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.font = Font(name="Cambria", color=COLOR_BLANCO, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name="Cambria", size=10, color=COLOR_TEXTO)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for col in range(1, ws.max_column + 1):
        letra = get_column_letter(col)
        ws.column_dimensions[letra].width = 32

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 62

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(ws.max_column)}{ws.max_row}"


# ============================================================
# 17. FIGURAS
# ============================================================

def extraer_irp(valor):
    match = re.search(r"IRP=(\d+\.\d+)", str(valor))
    if match:
        return float(match.group(1))
    return 0.0


def preparar_df_figuras(tabla):
    df = tabla.copy()
    df["IRP_2030"] = df["2030"].apply(extraer_irp)
    df["IRP_2035"] = df["2035"].apply(extraer_irp)
    df["IRP_2040"] = df["2040"].apply(extraer_irp)
    df["Cambio_2030_2040"] = df["IRP_2040"] - df["IRP_2030"]
    df["Promedio_IRP"] = df[["IRP_2030", "IRP_2035", "IRP_2040"]].mean(axis=1)
    return df


def color_por_irp(valor):
    if valor >= 0.78:
        return COLOR_ROJO_FIG
    if valor >= 0.62:
        return COLOR_VERDE_FIG
    if valor >= 0.48:
        return COLOR_NARANJA
    return COLOR_GRIS_FIG


def generar_mapa_calor(tabla):
    df = preparar_df_figuras(tabla)
    heat = df.sort_values("Promedio_IRP", ascending=False).head(10)

    matriz = heat.set_index("Tendencia identificada")[["IRP_2030", "IRP_2035", "IRP_2040"]]
    matriz.columns = ["2030", "2035", "2040"]

    fig, ax = plt.subplots(figsize=(18, 10))

    im = ax.imshow(matriz.values, aspect="auto", cmap="YlOrRd", vmin=0, vmax=1)

    ax.set_xticks(range(len(matriz.columns)))
    ax.set_xticklabels(matriz.columns, fontsize=13, weight="bold")

    etiquetas_y = [
        envolver_texto(recortar_texto(x, 92), 42)
        for x in matriz.index
    ]

    ax.set_yticks(range(len(matriz.index)))
    ax.set_yticklabels(etiquetas_y, fontsize=9)

    ax.set_xlabel("Año horizonte")
    ax.set_ylabel("Tendencia científica identificada")

    ax.set_title(
        "ATEFIC · Mapa de calor prospectivo sobre subcorpus orientado por reto",
        fontsize=18,
        weight="bold",
        color=COLOR_AZUL_OSCURO,
        pad=18
    )

    for i in range(matriz.shape[0]):
        for j in range(matriz.shape[1]):
            valor = matriz.iloc[i, j]
            etiqueta = etiqueta_corta(valor)
            color_texto = "white" if valor >= 0.62 else COLOR_AZUL_OSCURO

            ax.text(
                j,
                i,
                f"{etiqueta}\nIRP={valor:.2f}",
                ha="center",
                va="center",
                fontsize=9,
                color=color_texto,
                weight="bold"
            )

    ax.set_xticks(np.arange(-0.5, matriz.shape[1], 1), minor=True)
    ax.set_yticks(np.arange(-0.5, matriz.shape[0], 1), minor=True)
    ax.grid(which="minor", color="white", linestyle="-", linewidth=1.2)
    ax.tick_params(which="minor", bottom=False, left=False)

    cbar = fig.colorbar(im, ax=ax, fraction=0.032, pad=0.035)
    cbar.set_label("Índice de Relevancia Prospectiva", fontsize=10)

    fig.text(
        0.12,
        0.045,
        "Convención: Crítica ≥ 0.78 | Alta ≥ 0.62 | Consolidada ≥ 0.48 | Vigilancia < 0.48",
        fontsize=10,
        color=COLOR_AZUL_OSCURO,
        weight="bold"
    )

    fig.text(
        0.12,
        0.020,
        "Nota: primero se seleccionan documentos afines al reto y luego se identifican tendencias con TF-IDF, K-Means e IRP.",
        fontsize=8.5,
        color="#475569"
    )

    plt.tight_layout(rect=[0.04, 0.08, 0.94, 0.95])

    ruta = OUTPUT_DIR / "ATEFIC_01_mapa_calor_prospectivo_ML.png"
    plt.savefig(ruta, dpi=300, bbox_inches="tight")
    plt.close()

    return ruta


def generar_ranking(tabla):
    df = preparar_df_figuras(tabla)
    top = df.sort_values("IRP_2030", ascending=True).tail(10)

    fig, ax = plt.subplots(figsize=(18, 10))

    y = np.arange(len(top))
    colores = [color_por_irp(v) for v in top["IRP_2030"]]

    ax.barh(y, top["IRP_2030"], color=colores, edgecolor=COLOR_AZUL_OSCURO)

    ax.set_yticks(y)
    ax.set_yticklabels(
        [envolver_texto(recortar_texto(x, 100), 50) for x in top["Tendencia identificada"]],
        fontsize=9
    )

    ax.set_xlabel("Índice de Relevancia Prospectiva 2030")
    ax.set_title(
        "ATEFIC · Priorización de tendencias hacia 2030",
        fontsize=17,
        weight="bold",
        color=COLOR_AZUL_OSCURO
    )

    ax.set_xlim(0, max(1.0, top["IRP_2030"].max() + 0.08))
    ax.grid(axis="x", linestyle="--", alpha=0.25)

    for i, (_, row) in enumerate(top.iterrows()):
        ax.text(
            row["IRP_2030"] + 0.01,
            i,
            f"IRP={row['IRP_2030']:.2f}",
            va="center",
            fontsize=9,
            color=COLOR_AZUL_OSCURO,
            weight="bold"
        )

    plt.tight_layout()

    ruta = OUTPUT_DIR / "ATEFIC_02_ranking_IRP_2030.png"
    plt.savefig(ruta, dpi=300, bbox_inches="tight")
    plt.close()

    return ruta


def generar_lectura(tabla):
    df = preparar_df_figuras(tabla)
    top = df.sort_values("Promedio_IRP", ascending=False).head(10)

    fig, ax = plt.subplots(figsize=(16, 10))
    ax.axis("off")

    ax.text(
        0.03,
        0.96,
        "ATEFIC · Lectura prospectiva de tendencias 2030–2040",
        fontsize=18,
        weight="bold",
        color=COLOR_AZUL_OSCURO,
        ha="left",
        va="top"
    )

    ax.text(
        0.03,
        0.91,
        "Síntesis interpretativa del cambio proyectado del Índice de Relevancia Prospectiva.",
        fontsize=11,
        color="#475569",
        ha="left",
        va="top"
    )

    y = 0.84

    for idx, (_, row) in enumerate(top.iterrows(), start=1):
        cambio = row["Cambio_2030_2040"]
        promedio = row["Promedio_IRP"]

        if cambio > 0.05:
            direccion = "gana relevancia hacia 2040"
        elif cambio < -0.05:
            direccion = "pierde intensidad relativa hacia 2040"
        else:
            direccion = "permanece relativamente estable"

        tendencia = envolver_texto(row["Tendencia identificada"], 105)

        texto = (
            f"{idx}. {tendencia}\n"
            f"   Promedio IRP={promedio:.2f}; cambio 2030–2040={cambio:+.2f}; {direccion}."
        )

        ax.text(
            0.05,
            y,
            texto,
            fontsize=10,
            color=COLOR_TEXTO_FIG,
            ha="left",
            va="top",
            linespacing=1.3
        )

        y -= 0.075

    ax.text(
        0.03,
        0.06,
        "Nota: esta lectura se deriva del subcorpus seleccionado según el reto ingresado.",
        fontsize=9,
        color="#64748B",
        ha="left"
    )

    plt.tight_layout()

    ruta = OUTPUT_DIR / "ATEFIC_03_lectura_prospectiva_2030_2040.png"
    plt.savefig(ruta, dpi=300, bbox_inches="tight")
    plt.close()

    return ruta


def generar_matriz_cambio(tabla):
    df = preparar_df_figuras(tabla)
    plot = df.sort_values("Promedio_IRP", ascending=False).head(10).copy()
    plot["Orden"] = range(1, len(plot) + 1)

    fig, ax = plt.subplots(figsize=(17, 10))

    tamanos = 600 + plot["Promedio_IRP"] * 2000
    colores = [color_por_irp(v) for v in plot["IRP_2040"]]

    ax.scatter(
        plot["IRP_2030"],
        plot["Cambio_2030_2040"],
        s=tamanos,
        c=colores,
        alpha=0.78,
        edgecolor=COLOR_AZUL_OSCURO,
        linewidth=0.8
    )

    for _, row in plot.iterrows():
        ax.text(
            row["IRP_2030"],
            row["Cambio_2030_2040"],
            str(row["Orden"]),
            ha="center",
            va="center",
            fontsize=11,
            weight="bold",
            color="white"
        )

    ax.axhline(0, color=COLOR_GRIS_FIG, linestyle="--", alpha=0.45)
    ax.axvline(plot["IRP_2030"].median(), color=COLOR_GRIS_FIG, linestyle="--", alpha=0.35)

    ax.set_xlabel("IRP 2030")
    ax.set_ylabel("Cambio proyectado del IRP entre 2030 y 2040")
    ax.set_title(
        "ATEFIC · Matriz de cambio prospectivo 2030–2040",
        fontsize=17,
        weight="bold",
        color=COLOR_AZUL_OSCURO
    )

    ax.grid(alpha=0.18)

    leyenda = "Leyenda de tendencias:\n"
    for _, row in plot.iterrows():
        leyenda += f"{row['Orden']}. {recortar_texto(row['Tendencia identificada'], 70)}\n"

    ax.text(
        1.02,
        0.98,
        leyenda,
        transform=ax.transAxes,
        fontsize=8.5,
        va="top",
        ha="left",
        bbox=dict(boxstyle="round,pad=0.5", fc="white", ec="#CBD5E1", alpha=0.95)
    )

    fig.text(
        0.10,
        0.03,
        "Lectura: las burbujas con mayor altura ganan relevancia hacia 2040; las ubicadas a la derecha ya son relevantes en 2030.",
        fontsize=9,
        color="#475569"
    )

    plt.tight_layout(rect=[0.03, 0.06, 0.82, 0.95])

    ruta = OUTPUT_DIR / "ATEFIC_04_matriz_cambio_IRP_2030_2040.png"
    plt.savefig(ruta, dpi=300, bbox_inches="tight")
    plt.close()

    return ruta


def generar_figuras(tabla):
    generar_mapa_calor(tabla)
    generar_ranking(tabla)
    generar_lectura(tabla)
    generar_matriz_cambio(tabla)


# ============================================================
# 18. EJECUCIÓN GENERAL
# ============================================================

def ejecutar_analisis(reto):
    df_maestro, ruta_bib = leer_bibtex()

    df_subcorpus, info_subcorpus = filtrar_documentos_por_reto(df_maestro, reto)

    df_clusterizado, clusters, parametros_ml = construir_clusters(df_subcorpus, reto)

    df_scores = construir_proyeccion(df_clusterizado, clusters)

    tabla_tendencias = construir_tabla_tendencias(clusters, df_scores)

    tabla_subcorpus = construir_tabla_subcorpus(df_clusterizado)

    parametros = {
        **parametros_ml,
        **info_subcorpus
    }

    exportar_excel(
        tabla_tendencias=tabla_tendencias,
        tabla_subcorpus=tabla_subcorpus,
        reto=reto,
        ruta_bib=ruta_bib,
        parametros=parametros,
        info_subcorpus=info_subcorpus
    )

    # generar_figuras(tabla_tendencias)

    resumen = ""
    resumen += "METODOLOGÍA APLICADA\n"
    resumen += "La herramienta construyó un subcorpus orientado por el reto antes de identificar tendencias.\n"
    resumen += f"Documentos en corpus maestro: {info_subcorpus['documentos_corpus_maestro']}\n"
    resumen += f"Documentos seleccionados para este reto: {info_subcorpus['documentos_subcorpus']}\n"
    resumen += f"Porcentaje usado del corpus: {info_subcorpus['porcentaje_subcorpus']}%\n"
    resumen += f"Score promedio del subcorpus: {info_subcorpus['score_promedio']}\n\n"

    resumen += "TENDENCIAS PRINCIPALES\n"

    for _, row in tabla_tendencias.head(10).iterrows():
        resumen += (
            f"- {row['Tendencia identificada']} | "
            f"Impacto: {row['Nivel de impacto']} | "
            f"Velocidad: {row['Velocidad de evolución']} | "
            f"Afinidad con el reto: {row['Afinidad con el reto']} | "
            f"Score documental: {row['Score documental del subcorpus']} | "
            f"2030: {row['2030']} | "
            f"2035: {row['2035']} | "
            f"2040: {row['2040']}\n"
        )

    return resumen